# encoding: utf-8
#!usr/bin/python
__author__ = 'admin'

from openpyxl import workbook
from openpyxl import worksheet
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from common.Log import MyLog
import os
import sys

class OperateExcel:

    def __init__(self, sheetName):
        curPath = os.path.abspath(os.path.dirname(__file__))
        rootPath = curPath[:curPath.find("InterfaceTestDemo\\")+len("InterfaceTestDemo\\")]
        #print(rootPath)
        datafile = os.path.join(rootPath, "testdata\\test_cases.xlsx")
        #print(datafile)

        self.wb = load_workbook(datafile)
        #print(self.wb.sheetnames)

        self.ws = self.wb.get_sheet_by_name(sheetName)
        self.log = MyLog.get_log()


    def getUrlBySheet(self):
        self.log.logger.info("start read request url...")
        return self.ws.cell(1,2).value

    def getMethodBySheet(self):
        self.log.logger.info("start read request method...")
        return self.ws.cell(2,2).value

    def getParamsBySheet(self):
        self.log.logger.info("start read request parameters...")
        return self.ws.cell(3,2).value

    def getHeaderBySheet(self):
        self.log.logger.info("start read request header info...")
        return self.ws.cell(4,2).value

    def getBodyBySheet(self):
        self.log.logger.info("start read request body info...")
        return self.ws.cell(6,1).value

    def getExpectBySheet(self):
        self.log.logger.info("start read expect result...")
        return self.ws.cell(6,2).value



if __name__ == '__main__':
    print(OperateExcel('request_QD').getUrlBySheet())
    #print(OperateExcel('request_QD').getMethodBySheet())
    #print(OperateExcel('request_QD').getParamsBySheet())
    #print(OperateExcel('request_QD').getHeaderBySheet())
    #print(OperateExcel('request_QD').getBodyBySheet())
    #print(OperateExcel('request_QD').getExpectBySheet())
